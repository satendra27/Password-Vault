from tkinter import *
import tkinter as tk
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib
from tkinter import messagebox

root=Tk()
root.title("Password Vault")
root.geometry("800x600")
root.resizable(0,0)


file=pathlib.Path("Passwords.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Application Name"
    sheet["B1"]="Application URL"
    sheet["C1"]="Password"
    
    file.save("Passwords.xlsx")
    
def show():
    if password_entry.cget('show')=='*':
        password_entry.config(show='')
    else:
        password_entry.config(show='*')   
        
def submit():
    y=website_entry.get()
    z=email_entry.get()
    z1=password_entry.get()
    print(y)
    print(z)
    print(z1)
    
    file=openpyxl.load_workbook("Passwords.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    sheet.cell(column=3,row=sheet.max_row,value=z1)
    
    file.save("Passwords.xlsx")
    messagebox.showinfo("Password Vault","Your Data Have been Successfully Saved...")
    
    
frame=LabelFrame(root,text='Password Details',background="skyblue").pack(expand='yes',fill='both')


img=PhotoImage(file="output-onlinepngtools.png")
label_img=Label(frame,image=img,background="skyblue")
label_img.place(x=300,y=100)

label=Label(frame,text="Password Vault",font=("Imprint MT shadow",25),background="skyblue")
label.place(x=300,y=300)

website=Label(frame,text="Application Name or Website : ",background="skyblue",font=("cambria",13))
website.place(x=100,y=360)
email=Label(frame,text="Email/Username : ",background="skyblue",font=("cambria",13))
email.place(x=150,y=400)
password=Label(frame,text="Password : ",background="skyblue",font=("cambria",13))
password.place(x=150,y=440)

website_entry=Entry(frame,width=28,font=("arial",15),bd=2,highlightcolor="Blue",highlightthickness=2)
website_entry.place(x=350,y=360)
email_entry=Entry(frame,width=28,font=("arial",15),bd=2,highlightcolor="Blue",highlightthickness=2)
email_entry.place(x=350,y=400)
password_entry=Entry(frame,width=22,show='*',font=("arial",15),bd=2,highlightcolor="Blue",highlightthickness=2)
password_entry.place(x=350,y=440)


generate_btn=Button(frame,text="Show",activebackground="green",command=show,bd=3,highlightcolor="Blue",highlightthickness=2)
generate_btn.place(x=612,y=441)
add_btn=Button(frame,text="Add",width=36,command=submit,font=("arial",10),activebackground="lightblue",bd=3)
add_btn.place(x=350,y=490)

root.mainloop()
